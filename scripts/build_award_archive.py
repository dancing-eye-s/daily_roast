#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import json
import re
import subprocess
import sys
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
ARCHIVE_DIR = ROOT / "archive"
SEEDS_DIR = ARCHIVE_DIR / "seeds"
RAW_DIR = ARCHIVE_DIR / "raw"
NORMALIZED_DIR = ARCHIVE_DIR / "normalized"
REPORTS_DIR = ARCHIVE_DIR / "reports"
FRONTEND_DATA = ROOT / "data" / "archive.js"
JSON_OUTPUT = NORMALIZED_DIR / "award_archive.json"
SUMMARY_OUTPUT = REPORTS_DIR / "build_summary.json"
OCR_SUMMARY_OUTPUT = ROOT / "archive" / "ocr" / "summary.json"
OCR_PARSED_OUTPUT = ROOT / "archive" / "ocr" / "parsed_candidates.json"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/537.36 (KHTML, like Gecko) "
    "Chrome/123.0.0.0 Safari/537.36"
)

PDF_BACKEND = None
try:
    from pypdf import PdfReader  # type: ignore

    PDF_BACKEND = "pypdf"
except Exception:
    PdfReader = None


@dataclass
class SourceDocument:
    url: str
    kind: str
    title: str
    text: str
    links: list[str]
    fetched_at: str
    cache_path: str


AWARD_LEVELS = {"대상", "금상", "은상", "동상", "특별상"}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_dirs() -> None:
    for directory in [RAW_DIR, NORMALIZED_DIR, REPORTS_DIR]:
      directory.mkdir(parents=True, exist_ok=True)


def load_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def dump_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=2)
        handle.write("\n")


def slugify(value: str) -> str:
    value = value.strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    return value.strip("-") or "source"


def source_id(url: str, explicit_id: str | None = None) -> str:
    if explicit_id:
        return explicit_id
    parsed = urlparse(url)
    digest = hashlib.sha1(url.encode("utf-8")).hexdigest()[:10]
    stem = slugify(parsed.netloc + "-" + parsed.path)
    return f"{stem}-{digest}"


def cache_paths(url: str, explicit_id: str | None = None) -> tuple[Path, Path]:
    parsed = urlparse(url)
    cache_id = source_id(url, explicit_id)
    suffix = Path(parsed.path).suffix or ".html"
    raw_path = RAW_DIR / f"{cache_id}{suffix}"
    meta_path = RAW_DIR / f"{cache_id}.meta.json"
    return raw_path, meta_path


def fetch_url(url: str, force: bool = False, explicit_id: str | None = None) -> tuple[Path, Path]:
    raw_path, meta_path = cache_paths(url, explicit_id)
    if raw_path.exists() and meta_path.exists() and not force:
        return raw_path, meta_path

    response = requests.get(
        url,
        headers={"User-Agent": USER_AGENT},
        timeout=30,
    )
    response.raise_for_status()
    raw_path.parent.mkdir(parents=True, exist_ok=True)
    raw_path.write_bytes(response.content)
    dump_json(
        meta_path,
        {
            "url": url,
            "status_code": response.status_code,
            "content_type": response.headers.get("content-type", ""),
            "fetched_at": utc_now(),
            "size": len(response.content),
        },
    )
    return raw_path, meta_path


def normalize_whitespace(value: str) -> str:
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def parse_html(raw_bytes: bytes, url: str, kind: str, meta: dict[str, Any], cache_path: Path) -> SourceDocument:
    html = raw_bytes.decode("utf-8", errors="replace")
    soup = BeautifulSoup(html, "html.parser")
    title = ""
    if soup.title and soup.title.string:
        title = normalize_whitespace(soup.title.string)
    if not title:
        title = normalize_whitespace(soup.get_text(" ", strip=True)[:120])

    for tag in soup(["script", "style", "noscript"]):
        tag.decompose()

    body_text = normalize_whitespace(soup.get_text("\n", strip=True))
    links = sorted(
        {
            a.get("href")
            for a in soup.select("a[href]")
            if a.get("href") and a.get("href").startswith("http")
        }
    )
    return SourceDocument(
        url=url,
        kind=kind,
        title=title,
        text=body_text,
        links=links,
        fetched_at=meta.get("fetched_at", ""),
        cache_path=str(cache_path.relative_to(ROOT)),
    )


def parse_pdf(raw_bytes: bytes, url: str, kind: str, meta: dict[str, Any], cache_path: Path) -> SourceDocument:
    if PdfReader is None:
        text = "PDF parsing unavailable. Install pypdf to enable text extraction."
        title = Path(urlparse(url).path).name
        return SourceDocument(
            url=url,
            kind=kind,
            title=title,
            text=text,
            links=[],
            fetched_at=meta.get("fetched_at", ""),
            cache_path=str(cache_path.relative_to(ROOT)),
        )

    reader = PdfReader(str(cache_path))
    chunks: list[str] = []
    for page in reader.pages:
        try:
            page_text = page.extract_text() or ""
        except Exception:
            page_text = ""
        if page_text:
            chunks.append(page_text)

    title = Path(urlparse(url).path).name
    text = normalize_whitespace("\n".join(chunks))
    return SourceDocument(
        url=url,
        kind=kind,
        title=title,
        text=text,
        links=[],
        fetched_at=meta.get("fetched_at", ""),
        cache_path=str(cache_path.relative_to(ROOT)),
    )


def parse_xlsx(raw_bytes: bytes, url: str, kind: str, meta: dict[str, Any], cache_path: Path) -> SourceDocument:
    workbook = load_workbook(str(cache_path), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in workbook.worksheets:
        lines.append(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            values = [normalize_whitespace(str(value)) for value in row if value is not None and str(value).strip()]
            if values:
                lines.append(" | ".join(values))

    title = Path(urlparse(url).path).name
    return SourceDocument(
        url=url,
        kind=kind,
        title=title,
        text="\n".join(lines),
        links=[],
        fetched_at=meta.get("fetched_at", ""),
        cache_path=str(cache_path.relative_to(ROOT)),
    )


def parse_source(url: str, kind: str, force: bool = False, explicit_id: str | None = None) -> SourceDocument:
    raw_path, meta_path = fetch_url(url, force=force, explicit_id=explicit_id)
    meta = load_json(meta_path)
    raw_bytes = raw_path.read_bytes()
    if raw_path.suffix.lower() == ".pdf":
        return parse_pdf(raw_bytes, url, kind, meta, raw_path)
    if raw_path.suffix.lower() == ".xlsx":
        return parse_xlsx(raw_bytes, url, kind, meta, raw_path)
    return parse_html(raw_bytes, url, kind, meta, raw_path)


def collect_all_source_specs() -> list[dict[str, Any]]:
    source_index = load_json(SEEDS_DIR / "source_index.json")
    entry_seeds = load_json(SEEDS_DIR / "entries.seed.json")

    specs: dict[str, dict[str, Any]] = {}
    for source in source_index["sources"]:
        specs[source["url"]] = source
    for entry in entry_seeds:
        url = entry["sourceUrl"]
        if url not in specs:
            specs[url] = {
                "id": source_id(url),
                "url": url,
                "kind": infer_kind(url),
                "year": entry["year"],
                "role": "entry-source",
            }
    return list(specs.values())


def infer_kind(url: str) -> str:
    if "tvcf.co.kr" in url:
        return "tvcf-page"
    if url.lower().endswith(".xlsx"):
        return "award-table-xlsx"
    if url.lower().endswith(".pdf"):
        return "pdf"
    return "article"


def slug_fragment(value: str) -> str:
    fragment = re.sub(r"[^0-9a-zA-Z가-힣]+", "-", value.strip().lower())
    fragment = fragment.strip("-")
    return fragment[:80] or "entry"


def normalize_key(*values: str | int) -> str:
    return "|".join(normalize_whitespace(str(value)).lower() for value in values)


def make_auto_record(
    *,
    year: int,
    award_level: str,
    award_category: str,
    advertiser: str,
    campaign: str,
    source_type: str,
    source_url: str,
    source_document_title: str,
    source_fetched_at: str,
    source_cache_path: str,
    detail: str = "",
) -> dict[str, Any]:
    campaign = normalize_whitespace(campaign)
    advertiser = normalize_whitespace(advertiser)
    return {
        "id": f"auto-{year}-{slug_fragment(advertiser)}-{slug_fragment(campaign)}-{slug_fragment(award_level)}",
        "year": year,
        "awardLevel": award_level,
        "awardCategory": award_category,
        "brand": advertiser,
        "campaign": campaign,
        "copy": campaign,
        "detail": detail or "첨부 수상표 기반 자동 추출",
        "sourceType": source_type,
        "sourceUrl": source_url,
        "copyStatus": "inferred",
        "sourceDocumentTitle": source_document_title,
        "sourceFetchedAt": source_fetched_at,
        "sourceCachePath": source_cache_path,
    }


def extract_2014_xlsx_records(spec: dict[str, Any], doc: SourceDocument) -> list[dict[str, Any]]:
    workbook = load_workbook(ROOT / doc.cache_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    records: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=5, values_only=True):
        section = normalize_whitespace(str(row[0] or ""))
        award_level = normalize_whitespace(str(row[1] or ""))
        sub_section = normalize_whitespace(str(row[2] or ""))
        advertiser = normalize_whitespace(str(row[3] or ""))
        brand = normalize_whitespace(str(row[4] or ""))
        work_title = normalize_whitespace(str(row[5] or ""))
        if award_level not in AWARD_LEVELS or not work_title or advertiser == "수상작 없음":
            continue
        category = " / ".join(part for part in [section, sub_section] if part)
        campaign = work_title
        records.append(
            make_auto_record(
                year=spec["year"],
                award_level=award_level,
                award_category=category,
                advertiser=brand or advertiser,
                campaign=campaign,
                source_type="최종 수상작 XLSX",
                source_url=spec["url"],
                source_document_title=doc.title,
                source_fetched_at=doc.fetched_at,
                source_cache_path=doc.cache_path,
                detail=f"{advertiser} / {brand}".strip(" /"),
            )
        )
    return records


def extract_title_from_tail(tokens: list[str]) -> str:
    if not tokens:
        return ""
    generic_prefixes = {"기업", "pr", "pr캠페인", "기업pr", "브랜드", "ai", "3개월", "꽃병소화기"}
    title_tokens: list[str] = []
    idx = len(tokens) - 1
    base_consumed = False
    while idx >= 0:
        token = normalize_whitespace(tokens[idx])
        if not token:
            idx -= 1
            continue
        title_tokens.insert(0, token)
        idx -= 1
        if not base_consumed and not token.startswith((",", "-", "(")):
            base_consumed = True
            while idx >= 0:
                prev = normalize_whitespace(tokens[idx])
                prev_key = prev.lower()
                if prev.startswith((",", "-", "(")) or prev_key in generic_prefixes:
                    title_tokens.insert(0, prev)
                    idx -= 1
                    continue
                if len(title_tokens) == 1 and prev and re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9 .&+-]*", title_tokens[0]) and re.search(r"[가-힣]", prev):
                    title_tokens.insert(0, prev)
                    idx -= 1
                break
            break
    return normalize_whitespace(" ".join(title_tokens))


def extract_2018_pdf_records(spec: dict[str, Any], doc: SourceDocument) -> list[dict[str, Any]]:
    reader = PdfReader(str(ROOT / doc.cache_path))
    text = reader.pages[0].extract_text(extraction_mode="layout") or ""
    records: list[dict[str, Any]] = []
    current_category = ""
    for line in text.splitlines():
        if not re.match(r"^\s*\d+\s", line):
            continue
        parts = [part.strip() for part in re.split(r"\s{2,}", line.strip()) if part.strip()]
        if len(parts) < 3:
            continue

        idx = 1
        category_parts: list[str] = []
        while idx < len(parts) and parts[idx] not in AWARD_LEVELS:
            category_parts.append(parts[idx])
            idx += 1
        if idx >= len(parts):
            continue

        category_parts = [part for part in category_parts if part != "-"]
        if category_parts:
            current_category = " / ".join(category_parts)
        award_level = parts[idx]
        if award_level not in AWARD_LEVELS:
            continue
        rest = parts[idx + 1 :]
        if not rest:
            continue

        advertiser = normalize_whitespace(rest[0])
        if not advertiser or advertiser == "특별상":
            continue

        title = extract_title_from_tail(rest[1:])
        if not title:
            continue

        records.append(
            make_auto_record(
                year=spec["year"],
                award_level=award_level,
                award_category=current_category or "미분류",
                advertiser=advertiser,
                campaign=title,
                source_type="최종 수상작 PDF",
                source_url=spec["url"],
                source_document_title=doc.title,
                source_fetched_at=doc.fetched_at,
                source_cache_path=doc.cache_path,
                detail="PDF 표 기반 자동 추출",
            )
        )
    return records


def build_attachment_records(source_specs: list[dict[str, Any]], force_fetch: bool = False) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for spec in source_specs:
        if spec.get("role") != "award-table":
            continue
        doc = parse_source(spec["url"], spec["kind"], force=force_fetch, explicit_id=spec.get("id"))
        if spec["kind"] == "award-table-xlsx":
            records.extend(extract_2014_xlsx_records(spec, doc))
        elif spec["kind"] == "award-table-pdf":
            records.extend(extract_2018_pdf_records(spec, doc))
    return records


def fetch_all(force: bool = False) -> dict[str, Any]:
    ensure_dirs()
    fetched: list[dict[str, Any]] = []
    for spec in collect_all_source_specs():
        doc = parse_source(spec["url"], spec["kind"], force=force, explicit_id=spec.get("id"))
        fetched.append(
            {
                "id": spec.get("id"),
                "url": spec["url"],
                "kind": spec["kind"],
                "title": doc.title,
                "cache_path": doc.cache_path,
                "fetched_at": doc.fetched_at,
            }
        )
    report = {
        "generatedAt": utc_now(),
        "sourceCount": len(fetched),
        "sources": fetched,
    }
    dump_json(REPORTS_DIR / "fetch_report.json", report)
    return report


def build_archive(force_fetch: bool = False) -> dict[str, Any]:
    ensure_dirs()
    seeds = load_json(SEEDS_DIR / "entries.seed.json")
    source_specs = collect_all_source_specs()
    documents: dict[str, SourceDocument] = {}
    manual_records: list[dict[str, Any]] = []
    statuses: dict[str, int] = {}

    for seed in seeds:
        url = seed["sourceUrl"]
        if url not in documents:
            documents[url] = parse_source(
                url,
                seed.get("sourceKind", infer_kind(url)),
                force=force_fetch,
                explicit_id=seed.get("sourceId"),
            )

        doc = documents[url]
        statuses[seed["copyStatus"]] = statuses.get(seed["copyStatus"], 0) + 1
        record = {
            "id": seed["id"],
            "year": seed["year"],
            "awardLevel": seed["awardLevel"],
            "awardCategory": seed["awardCategory"],
            "brand": seed["brand"],
            "campaign": seed["campaign"],
            "copy": seed["copy"],
            "detail": seed.get("detail", ""),
            "sourceType": seed["sourceType"],
            "sourceUrl": url,
            "copyStatus": seed["copyStatus"],
            "sourceDocumentTitle": doc.title,
            "sourceFetchedAt": doc.fetched_at,
            "sourceCachePath": doc.cache_path,
        }
        if seed.get("notes"):
            record["notes"] = seed["notes"]
        manual_records.append(record)

    attachment_records = build_attachment_records(source_specs, force_fetch=force_fetch)
    existing_keys = {
        normalize_key(record["year"], record["awardLevel"], record["brand"], record["campaign"])
        for record in manual_records
    }
    records = list(manual_records)
    for record in attachment_records:
        key = normalize_key(record["year"], record["awardLevel"], record["brand"], record["campaign"])
        if key in existing_keys:
            continue
        records.append(record)
        statuses[record["copyStatus"]] = statuses.get(record["copyStatus"], 0) + 1
        existing_keys.add(key)

    records.sort(key=lambda item: (-item["year"], item["brand"], item["campaign"], item["awardLevel"]))
    payload = {
        "generatedAt": utc_now(),
        "recordCount": len(records),
        "copyStatusCounts": statuses,
        "pdfBackend": PDF_BACKEND or "disabled",
        "records": records,
    }
    dump_json(JSON_OUTPUT, payload)
    write_frontend_data(records)

    summary = {
        "generatedAt": payload["generatedAt"],
        "recordCount": len(records),
        "years": sorted({record["year"] for record in records}, reverse=True),
        "copyStatusCounts": statuses,
        "pdfBackend": payload["pdfBackend"],
        "output": str(JSON_OUTPUT.relative_to(ROOT)),
        "frontendOutput": str(FRONTEND_DATA.relative_to(ROOT)),
    }
    if OCR_SUMMARY_OUTPUT.exists():
        summary["ocrSummary"] = str(OCR_SUMMARY_OUTPUT.relative_to(ROOT))
    dump_json(SUMMARY_OUTPUT, summary)
    return summary


def run_ocr_pipeline(force_images: bool = False) -> dict[str, Any]:
    command = ["python3", str(ROOT / "scripts" / "build_ocr_artifacts.py")]
    if force_images:
        command.append("--force-images")
    result = subprocess.run(command, check=True, capture_output=True, text=True)
    return json.loads(result.stdout)


def run_ocr_parse_pipeline() -> dict[str, Any]:
    command = ["python3", str(ROOT / "scripts" / "parse_ocr_candidates.py")]
    result = subprocess.run(command, check=True, capture_output=True, text=True)
    return json.loads(result.stdout)


def write_frontend_data(records: list[dict[str, Any]]) -> None:
    FRONTEND_DATA.parent.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(records, ensure_ascii=False, indent=2)
    content = (
        "// Generated by scripts/build_award_archive.py\n"
        f"window.AD_ARCHIVE = {payload};\n"
    )
    FRONTEND_DATA.write_text(content, encoding="utf-8")


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Build and cache the Korea Ad Awards archive dataset."
    )
    parser.add_argument(
        "command",
        choices=["fetch", "build", "ocr", "all"],
        nargs="?",
        default="all",
        help="fetch raw sources, build normalized outputs, run OCR, or do all",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="re-fetch remote sources even when cached files already exist",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv or sys.argv[1:])
    try:
        if args.command in {"fetch", "all"}:
            report = fetch_all(force=args.force)
            print(
                json.dumps(
                    {
                        "step": "fetch",
                        "sourceCount": report["sourceCount"],
                        "generatedAt": report["generatedAt"],
                    },
                    ensure_ascii=False,
                )
            )
        if args.command in {"ocr", "all"}:
            ocr_summary = run_ocr_pipeline(force_images=args.force)
            print(json.dumps({"step": "ocr", **ocr_summary}, ensure_ascii=False))
            ocr_parse_summary = run_ocr_parse_pipeline()
            print(json.dumps({"step": "ocr-parse", **ocr_parse_summary}, ensure_ascii=False))
        if args.command in {"build", "all"}:
            summary = build_archive(force_fetch=args.force)
            print(json.dumps({"step": "build", **summary}, ensure_ascii=False))
        return 0
    except requests.HTTPError as exc:
        print(f"HTTP error while fetching sources: {exc}", file=sys.stderr)
        return 1
    except requests.RequestException as exc:
        print(f"Network error while fetching sources: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
